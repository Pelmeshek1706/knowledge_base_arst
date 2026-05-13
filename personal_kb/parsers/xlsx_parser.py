from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from personal_kb.core.errors import ParsingError
from personal_kb.parsers.base import BaseParser
from personal_kb.schemas.document import ParsedDocument

_INLINE_WHITESPACE_RE = re.compile(r"[^\S\n]+")


class XlsxParser(BaseParser):
    supported_extensions = ("xlsx",)

    def parse(self, file_path: Path, *, source_id: str | None = None) -> ParsedDocument:
        normalized_source_id = self._normalize_source_id(file_path, source_id)
        raw_text, blocks = self._extract_workbook(file_path, normalized_source_id)

        if not self._has_usable_text(raw_text):
            raise ParsingError(
                f"xlsx text extraction produced no usable text for {normalized_source_id}"
            )

        return ParsedDocument(
            source_id=normalized_source_id,
            file_path=normalized_source_id,
            file_name=file_path.name,
            file_extension="xlsx",
            title=file_path.stem,
            raw_text=raw_text,
            metadata=self._build_metadata(file_path),
            structured_blocks=blocks,
        )

    def _extract_workbook(
        self, file_path: Path, source_id: str
    ) -> tuple[str, list[dict[str, Any]]]:
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        try:
            raw_lines: list[str] = []
            blocks: list[dict[str, Any]] = []

            for sheet in workbook.worksheets:
                sheet_blocks = self._extract_sheet_blocks(sheet, source_id)
                if not sheet_blocks:
                    continue
                raw_lines.append(f"[Sheet] {sheet.title}")
                for block in sheet_blocks:
                    raw_lines.append(block["text"])
                raw_lines.append("")
                blocks.extend(sheet_blocks)

            return self._normalize_text("\n".join(raw_lines)), blocks
        finally:
            workbook.close()

    def _extract_sheet_blocks(self, sheet: Any, source_id: str) -> list[dict[str, Any]]:
        blocks: list[dict[str, Any]] = []
        current_rows: list[dict[str, Any]] = []

        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            row_values = [self._normalize_cell_value(cell.value) for cell in row]
            non_empty_columns = [cell.column for cell, value in zip(row, row_values) if value]
            if not non_empty_columns:
                if current_rows:
                    blocks.append(self._finalize_sheet_block(current_rows, sheet.title, source_id))
                    current_rows = []
                continue

            current_rows.append(
                {
                    "row_index": row_index,
                    "values": row_values,
                    "start_col": min(non_empty_columns),
                    "end_col": max(non_empty_columns),
                }
            )

        if current_rows:
            blocks.append(self._finalize_sheet_block(current_rows, sheet.title, source_id))

        return blocks

    def _finalize_sheet_block(
        self, rows: list[dict[str, Any]], sheet_name: str, source_id: str
    ) -> dict[str, Any]:
        start_row = rows[0]["row_index"]
        end_row = rows[-1]["row_index"]
        start_col = min(row["start_col"] for row in rows)
        end_col = max(row["end_col"] for row in rows)
        cell_range = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )
        text = self._rows_to_text(rows, start_col, end_col)
        return {
            "type": "xlsx_range",
            "sheet_name": sheet_name,
            "cell_range": cell_range,
            "start_row": start_row,
            "end_row": end_row,
            "start_col": start_col,
            "end_col": end_col,
            "text": text,
            "source_ref": {
                "file_path": source_id,
                "sheet": sheet_name,
                "cell_range": cell_range,
            },
        }

    def _rows_to_text(
        self,
        rows: list[dict[str, Any]],
        start_col: int,
        end_col: int,
    ) -> str:
        rendered_rows: list[str] = []
        for row in rows:
            values = row["values"]
            rendered_cells: list[str] = []
            for column_index in range(start_col, end_col + 1):
                value_index = column_index - 1
                rendered_cells.append(values[value_index] if value_index < len(values) else "")
            rendered_row = "\t".join(rendered_cells).rstrip()
            if rendered_row:
                rendered_rows.append(rendered_row)
        return "\n".join(rendered_rows).strip()

    def _normalize_cell_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return self._normalize_text(value)
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return self._normalize_text(str(value))

    def _normalize_text(self, text: str) -> str:
        normalized_lines: list[str] = []
        previous_blank = False

        for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            normalized_line = _INLINE_WHITESPACE_RE.sub(" ", line).strip()
            if not normalized_line:
                if previous_blank:
                    continue
                normalized_lines.append("")
                previous_blank = True
                continue

            normalized_lines.append(normalized_line)
            previous_blank = False

        return "\n".join(normalized_lines).strip()

    def _has_usable_text(self, text: str) -> bool:
        return bool(text.strip()) and bool(re.search(r"[A-Za-z0-9]", text))
